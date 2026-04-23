# src/clean_project/analysis/metrics.py
import pandas as pd
import unidecode
import numpy as np
from pathlib import Path
import clean_project.config.settings as config
import json
import hashlib
from datetime import datetime
from openpyxl.utils import get_column_letter

# ---------------------------------------------------
# PILARES PARA MÉTRICA DE ACEPTACIÓN GLOBAL
# ---------------------------------------------------
PILARES = [
    "legitimacion",
    "efectividad",
    "justicia_equidad",
    "confianza_institucional"
]

# ---------------------------------------------------
# FUNCIONES AUXILIARES DE CARGA
# ---------------------------------------------------
def inferir_red(nombre_archivo):
    nombre = nombre_archivo.lower()
    if "bluesky" in nombre: return "bluesky"
    if "reddit" in nombre: return "reddit"
    if "twitter" in nombre or "x_" in nombre: return "twitter"
    if "youtube" in nombre: return "youtube"
    if "linkedin" in nombre: return "linkedin"
    return nombre.replace("_global_dataset", "").replace(".csv", "")

def cargar_datos(data_path):
    archivos = list(data_path.glob("*_analizado.csv"))
    
    if not archivos:
        print(f"❌ No se encontraron archivos *_analizado.csv en {data_path}")
        return []

    all_rows = []
    for archivo in archivos:
        red = inferir_red(archivo.name)
        sep = "," 
        try:
            df = pd.read_csv(archivo, sep=sep, on_bad_lines='skip')
            if df.empty:
                print(f"⚠️ El archivo {archivo.name} está vacío. Saltando.")
                continue
            print(f"\n📥 Cargando archivo: {archivo.name} → Red detectada: {red} | Filas: {len(df)}")
            all_rows.append((red, df))
        except pd.errors.EmptyDataError:
            print(f"⚠️ El archivo {archivo.name} está vacío o corrupto. Saltando.")
        except Exception as e:
            print(f"⚠️ Error al leer {archivo.name}: {e}")
            
    return all_rows

def mapear_columnas_pilares(df):
    mapeo = {}
    columnas_normalizadas = df.columns.str.lower().map(lambda x: unidecode.unidecode(str(x)))

    variantes = {
        "legitimacion": ["sent_legitimación_sociopolítica", "sent_Legitimación_sociopolítica"],
        "efectividad": ["sent_efectividad_percibida", "sent_Efectividad_percibida"],
        "justicia_equidad": ["sent_justicia_y_equidad_percibida", "sent_Justicia_y_equidad_percibida"],
        "confianza_institucional": ["sent_confianza_y_legitimidad_institucional", "sent_Confianza_y_legitimidad_institucional"]
    }

    for pilar, posibles in variantes.items():
        for col in posibles:
            col_normalizada = unidecode.unidecode(col).lower()
            if col_normalizada in columnas_normalizadas.tolist():
                indice = columnas_normalizadas.tolist().index(col_normalizada)
                mapeo[pilar] = df.columns[indice]
                break
    return mapeo

# ---------------------------------------------------
# NUEVAS FUNCIONES PARA GENERAR EXCEL UNIFICADO
# ---------------------------------------------------
def estandarizar_para_excel(df, red, mapeo_pilares):
    """
    Transforma el DF de cualquier red al formato estándar solicitado para el Excel final.
    Limpiado: Sin Categoría, Sin Sentimiento General, Sin Explicaciones, Solo Tema 1.
    """
    nuevo_df = pd.DataFrame()

    # 1. ID ANONIMIZADO
    col_user = 'usuario_comentario' if red == 'youtube' else 'usuario'
    if col_user in df.columns:
        nuevo_df['ID'] = df[col_user].astype(str).fillna('unknown').apply(
            lambda x: hashlib.sha256(x.encode()).hexdigest()[:16]
        )
    else:
        nuevo_df['ID'] = 'unknown'

    # 2. FECHA Y HORA (ROBUSTO)
    col_date = 'fecha_comentario' if red == 'youtube' else 'fecha'
    fechas_list = []
    horas_list = []

    if col_date in df.columns:
        raw_dates = df[col_date].astype(str)
        # Intento con UTC
        fechas_dt = pd.to_datetime(raw_dates, utc=True, errors='coerce')

        for val_raw, dt_obj in zip(raw_dates, fechas_dt):
            if pd.notna(dt_obj):
                fechas_list.append(dt_obj.strftime('%Y-%m-%d'))
                horas_list.append(dt_obj.strftime('%H:%M'))
            else:
                # Plan B manual
                val_str = val_raw.strip()
                if 'T' in val_str:
                    parts = val_str.split('T')
                    fechas_list.append(parts[0])
                    hora_part = parts[1].split('+')[0].split('-')[0].split('Z')[0]
                    horas_list.append(hora_part[:5])
                elif ' ' in val_str:
                    parts = val_str.split(' ')
                    fechas_list.append(parts[0])
                    if len(parts) > 1 and ':' in parts[1]:
                        horas_list.append(parts[1][:5])
                    else:
                        horas_list.append("")
                else:
                    fechas_list.append(val_str)
                    horas_list.append("")
    else:
        fechas_list = [""] * len(df)
        horas_list = [""] * len(df)

    nuevo_df['FECHA'] = fechas_list
    nuevo_df['HRS'] = horas_list

    # 3. CONTENIDO TEXTUAL
    nuevo_df['TITULO'] = ""
    nuevo_df['DESCRIPCION'] = ""
    nuevo_df['CONTENIDO'] = ""

    if red == 'youtube':
        if 'titulo_video' in df.columns: nuevo_df['TITULO'] = df['titulo_video']
        if 'descripcion_video' in df.columns: nuevo_df['DESCRIPCION'] = df['descripcion_video']
        if 'contenido' in df.columns: nuevo_df['CONTENIDO'] = df['contenido']
    elif red == 'reddit':
        if 'post_title' in df.columns: nuevo_df['TITULO'] = df['post_title']
        if 'post_selftext' in df.columns: nuevo_df['DESCRIPCION'] = df['post_selftext']
        if 'contenido' in df.columns: nuevo_df['CONTENIDO'] = df['contenido']
    else:
        if 'contenido' in df.columns: nuevo_df['CONTENIDO'] = df['contenido']

    # 4. METADATOS (Sin Categoría)
    cols_meta = {
        'TERMINO DE BUSQUEDA': ['search_keyword'],# termino_busqueda', 'query', 'search_term'],
        'IDIOMA': ['keyword_languages']# idioma', 'lang', 'language']
    }
    for col_final, posibles in cols_meta.items():
        nuevo_df[col_final] = ""
        for posible in posibles:
            if posible in df.columns:
                nuevo_df[col_final] = df[posible]
                break

    # 5. PILARES (Renombrado a siglas)
    map_nombres_pilares = {
        "legitimacion": "SENTIMIENTO_LS",
        "efectividad": "SENTIMIENTO_E",
        "justicia_equidad": "SENTIMIENTO_JE",
        "confianza_institucional": "SENTIMIENTO_CL"
    }

    for pilar_key, col_original in mapeo_pilares.items():
        nombre_final = map_nombres_pilares.get(pilar_key, f"SENTIMIENTO_{pilar_key.upper()}")
        nuevo_df[nombre_final] = df[col_original]

    # Rellenar pilares faltantes
    for pilar_key, nombre_final in map_nombres_pilares.items():
        if nombre_final not in nuevo_df.columns:
            nuevo_df[nombre_final] = ""

    # 6. SOLO TEMA 1 Y SENTIMIENTO TEMA 1
    # Buscamos específicamente Topic_1 (insensible a mayúsculas)
    col_t1 = next((c for c in df.columns if c.lower() == 'topic_1'), None)
    nuevo_df['TEMA_1'] = df[col_t1] if col_t1 else ""

    # Buscamos Sentimiento para Topic 1
    # Puede llamarse Sentimiento_Topic_1, Sentimiento_1, etc.
    posibles_s1 = ['sentimiento_topic_1', 'sentimiento_1', 'sent_topic_1']
    col_s1 = next((c for c in df.columns if c.lower() in posibles_s1), None)
    nuevo_df['SENTIMIENTO_TEMA_1'] = df[col_s1] if col_s1 else ""

    # Añadir Red Social
    nuevo_df['RED_SOCIAL'] = red
    
    # 7. ORDEN FINAL DE COLUMNAS
    cols_finales = [
        'ID', 'FECHA', 'HRS', 'RED_SOCIAL', 
        'TITULO', 'DESCRIPCION', 'CONTENIDO', 
        'TERMINO DE BUSQUEDA', 'IDIOMA', 
        'SENTIMIENTO_LS', 'SENTIMIENTO_E', 'SENTIMIENTO_JE', 'SENTIMIENTO_CL', 'SENTIMIENTO_M',
        'TEMA_1', 'SENTIMIENTO_TEMA_1'
    ]
    
    # Aseguramos que solo devolvemos estas columnas (rellenando si falta alguna por seguridad)
    for c in cols_finales:
        if c not in nuevo_df.columns:
            nuevo_df[c] = ""
            
    return nuevo_df[cols_finales]

def generar_datasets_finales(all_rows, output_dir):
    """
    Genera un ÚNICO Excel con dos pestañas:
    1. Resultados Utilizados
    2. Resultados Descartados
    Aplica formato de TABLA (Filtros y ordenación).
    """
    print("\n🛡️  Generando Excel Unificado Final...")

    lista_relevantes = []
    lista_descartados = []

    for red, df in all_rows:
        # 1. Detectar mapeo de pilares
        mapeo = mapear_columnas_pilares(df)
        
        # 2. Estandarizar DF (Limpieza aplicada aquí)
        df_std = estandarizar_para_excel(df, red, mapeo)
        
        # 3. Filtrar Relevancia
        cols_pilares_std = ["SENTIMIENTO_LS", "SENTIMIENTO_E", "SENTIMIENTO_JE", "SENTIMIENTO_CL", "SENTIMIENTO_M"]
        
        # Convertir a numérico para filtrar
        temp_vals = df_std[cols_pilares_std].apply(pd.to_numeric, errors='coerce').fillna(2)
        es_relevante = temp_vals.isin([1, -1]).any(axis=1)

        # 4. Separar
        df_rel = df_std[es_relevante].copy()
        df_desc = df_std[~es_relevante].copy()
        
        lista_relevantes.append(df_rel)
        lista_descartados.append(df_desc)

    # 5. Concatenar
    df_final_relevante = pd.concat(lista_relevantes, ignore_index=True) if lista_relevantes else pd.DataFrame()
    df_final_descartado = pd.concat(lista_descartados, ignore_index=True) if lista_descartados else pd.DataFrame()

    # 6. Guardar Excel con FORMATO
    output_file = output_dir / "DATASET_FINAL_UNIFICADO.xlsx"
    
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # --- Pestaña 1: Relevantes ---
            sheet_name_rel = 'Resultados Utilizados'
            df_final_relevante.to_excel(writer, sheet_name=sheet_name_rel, index=False)
            
            if not df_final_relevante.empty:
                worksheet = writer.sheets[sheet_name_rel]
                num_rows = len(df_final_relevante) + 1
                num_cols = len(df_final_relevante.columns)
                last_col_letter = get_column_letter(num_cols)
                
                # Filtros
                worksheet.auto_filter.ref = f"A1:{last_col_letter}{num_rows}"
                
                # Ancho de columnas
                for i, col in enumerate(df_final_relevante.columns):
                    col_letter = get_column_letter(i + 1)
                    worksheet.column_dimensions[col_letter].width = len(col) + 5

            # --- Pestaña 2: Descartados ---
            sheet_name_desc = 'Resultados Descartados'
            df_final_descartado.to_excel(writer, sheet_name=sheet_name_desc, index=False)
            
            if not df_final_descartado.empty:
                worksheet = writer.sheets[sheet_name_desc]
                num_rows = len(df_final_descartado) + 1
                num_cols = len(df_final_descartado.columns)
                last_col_letter = get_column_letter(num_cols)
                
                # Filtros
                worksheet.auto_filter.ref = f"A1:{last_col_letter}{num_rows}"
                
                # Ancho de columnas
                for i, col in enumerate(df_final_descartado.columns):
                    col_letter = get_column_letter(i + 1)
                    worksheet.column_dimensions[col_letter].width = len(col) + 5
        
        print(f"   ✅ Excel generado con filtros: {output_file.name}")
        print(f"      - Utilizados: {len(df_final_relevante)}")
        print(f"      - Descartados: {len(df_final_descartado)}")

    except Exception as e:
        print(f"   ❌ Error al guardar Excel: {e}")

# ---------------------------------------------------
# CÁLCULO DE MÉTRICAS (Sin cambios)
# ---------------------------------------------------
def aceptacion_global_promedio_pilares(all_rows, pilares):
    aceptacion_por_pilar = {}
    total_filas = 0
    detalles_por_red = {}

    for red, df in all_rows:
        mapeo = mapear_columnas_pilares(df)
        if not mapeo:
            detalles_por_red[red] = {"error": "No se detectaron columnas de pilares"}
            continue

        print(f"\n🔹 Procesando red: {red}")
        columnas_disponibles = list(mapeo.values())
        
        df_pilares = df[columnas_disponibles].copy()
        df_pilares = df_pilares.apply(pd.to_numeric, errors='coerce').fillna(2)

        mask_relevantes = df_pilares.isin([1, -1]).any(axis=1)
        df_pilares = df_pilares[mask_relevantes]
        n_filas = len(df_pilares)
        print(f"   Filas relevantes para cálculo: {n_filas}")
        total_filas += n_filas
        
        stats_red = {
            "filas_relevantes": n_filas,
            "conteos": {}
        }
        
        for p in pilares:
            if p in mapeo:
                vals = df_pilares[mapeo[p]]
                vals = vals[vals != 2]
                pos = (vals == 1).sum()
                neg = (vals == -1).sum()
                neu = (vals == 0).sum()
                denom = pos + neg + neu
                
                aceptacion_por_pilar.setdefault(p, []).append(float((pos - neg) / denom) if denom > 0 else np.nan)
                stats_red["conteos"][p] = {"pos": int(pos), "neg": int(neg), "neu": int(neu)}
            else:
                stats_red["conteos"][p] = {"pos": 0, "neg": 0, "neu": 0}

        detalles_por_red[red] = stats_red    

    promedio_pilares = {}
    for p, lista in aceptacion_por_pilar.items():
        valores_validos = [v for v in lista if not pd.isna(v)]
        promedio_pilares[p] = float(np.mean(valores_validos)) if valores_validos else np.nan

    valores_validos_global = [v for v in promedio_pilares.values() if not pd.isna(v)]
    aceptacion_global = float(np.mean(valores_validos_global)) if valores_validos_global else np.nan
    aceptacion_global_norm = (aceptacion_global + 1) / 2 * 100 if not pd.isna(aceptacion_global) else np.nan

    return {
        "aceptacion_por_pilar": promedio_pilares,
        "aceptacion_global": aceptacion_global,
        "aceptacion_global_norm": aceptacion_global_norm,
        "n_publicaciones_usadas": total_filas,
        "detalles_por_red": detalles_por_red
    }

def interpretar_aceptacion(valor_norm):
    if pd.isna(valor_norm): return "Sin datos"
    if valor_norm <= 20: return "Muy baja aceptación"
    elif valor_norm <= 40: return "Baja aceptación"
    elif valor_norm <= 60: return "Aceptación media"
    elif valor_norm <= 80: return "Alta aceptación"
    else: return "Muy alta aceptación"
    
def generar_informe(resultados, all_rows, pilares):
    n_descartados = sum(df.shape[0] - df[df.isin([1,-1]).any(axis=1)].shape[0] for _, df in all_rows)
    
    if resultados['aceptacion_por_pilar']:
        pilar_mas_influyente = max(resultados['aceptacion_por_pilar'], key=lambda k: abs(resultados['aceptacion_por_pilar'][k] if not pd.isna(resultados['aceptacion_por_pilar'][k]) else 0))
    else:
        pilar_mas_influyente = "Ninguno"
    
    informe = {
        "Aceptación Global [%]": resultados["aceptacion_global_norm"],
        "Interpretación": interpretar_aceptacion(resultados["aceptacion_global_norm"]),
        "Pilar más influyente": pilar_mas_influyente,
        "Cantidad de publicaciones usadas": resultados["n_publicaciones_usadas"],
        "Mensajes descartados": n_descartados,
        "Aceptación por pilar": resultados["aceptacion_por_pilar"],
        "Detalle por Red": resultados["detalles_por_red"]
    }
    return informe    

# ---------------------------------------------------
# FUNCIÓN PRINCIPAL
# ---------------------------------------------------
def metrics(u_conf):
    data_path = Path(u_conf.general["output_folder"])
    all_rows = cargar_datos(data_path)
    if not all_rows:
        return

    # 1. Calcular métricas
    resultados = aceptacion_global_promedio_pilares(all_rows, PILARES)

    print("\n📘 MÉTRICA DE ACEPTACIÓN GLOBAL")
    print(f"   Aceptación Global [0-100]: {resultados['aceptacion_global_norm']:.2f}%")
    
    # 2. Generar informe
    informe = generar_informe(resultados, all_rows, PILARES)

    # 3. Guardar JSON y CSV
    output_dir = data_path
    output_dir.mkdir(exist_ok=True)

    informe_simple = informe.copy()
    del informe_simple["Detalle por Red"]

    json_path = output_dir / "aceptacion_global.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(informe, f, indent=4, ensure_ascii=False)

    csv_path = output_dir / "aceptacion_global.csv"
    pd.DataFrame([informe_simple]).to_csv(csv_path, index=False)

    # 4. Guardar TXT
    txt_path = output_dir / "aceptacion_global.txt"
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write("📘 INFORME DE ACEPTACIÓN GLOBAL\n")
        f.write("="*40 + "\n\n")
        for k, v in informe.items():
            if k == "Detalle por Red": continue
            if isinstance(v, float): f.write(f"{k:<35}: {v:.2f}\n")
            elif isinstance(v, dict):
                f.write(f"{k}:\n")
                for p, val in v.items(): f.write(f"   {p:<25}: {val:.2f}\n")
            else: f.write(f"{k:<35}: {v}\n")
        
        f.write("\n" + "="*40 + "\nDETALLE POR RED SOCIAL\n" + "="*40 + "\n")
        detalles = informe.get("Detalle por Red", {})
        for red, stats in detalles.items():
            f.write(f"\n🔹 Red: {red.upper()}\n")
            if "error" in stats: f.write(f"   ⚠️ {stats['error']}\n")
            else:
                f.write(f"   Filas relevantes: {stats['filas_relevantes']}\n")
                for pilar, counts in stats['conteos'].items():
                    f.write(f"      {pilar}: +{counts['pos']} / -{counts['neg']} / ={counts['neu']}\n")

    # 5. Generar Excel Final Unificado
    generar_datasets_finales(all_rows, output_dir)

    print(f"\n📁 Resultados guardados en:\n  - {txt_path}\n  - {json_path}\n  - {output_dir / 'DATASET_FINAL_UNIFICADO.xlsx'}")

